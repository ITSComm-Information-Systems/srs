from project.pinnmodels import UmEcommMbidWarehseInput, UmEcommMbidCriticalDate, UmEcommMbidVendorInput, UmEcommMbidCommodityV, UmEcommMbidVendorV, UmBomProcurementUsersV

# Url
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.shortcuts import render, redirect, reverse

# Functional
from apps.mbid.forms import createCycle
import math
import time
import threading
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from io import BytesIO
from collections import OrderedDict
from openpyxl.styles import Protection

from datetime import date, datetime
from django.core.mail import EmailMessage
from django.db import connections
from django.db.models import Q

# Security
from django.views.decorators.http import condition
from django.contrib.auth.decorators import login_required, permission_required

# Required information for most pages
convert = {'01': 'January', '02': 'February', '03': 'March', '04': 'April', '05': 'May', '06': 'June',
           '07': 'July', '08': 'August', '09': 'September', '10': 'October', '11': 'November', '12': 'December'}
reverse_convert = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06',
                   'July': '07', 'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12'}

# Call on views that display or use cycle information


def cycle():
    cycle_info = {'open': False, 'viewable': False}
    # current cycle exists: open=True, viewable=True
    # current cycle does not exist, no closed cycle exists: open=False, viewable=False
    # current cycle does not exist, closed cycle exists: open=False, viewable=True
    try: # seeing if open bidding cycle exists
        current_cycle = UmEcommMbidCriticalDate.objects.get(bidding_closed='O')
        # Need to auto-close the bidding cycle
        if current_cycle.bidding_close_time < datetime.now():
            UmEcommMbidCriticalDate.objects.filter(bidding_closed='O').update(bidding_closed='C')
            # current_cycle.bidding_closed = 'C'
            # current_cycle.save(update_fields=["bidding_closed"])
            cycle_info['open'] = False
            cycle_info['viewable'] = True
            print(cycle_info)
            return cycle_info
        # Bidding cycle is open
        else:
            cycle_info['current_year'] = current_cycle.bidding_year
            cycle_info['current_month_num'] = current_cycle.bidding_month
            cycle_info['current_month'] = convert[current_cycle.bidding_month]

            cycle_info['current_open_date'] = datetime.strptime(
                str(current_cycle.bidding_open_date), '%Y-%m-%d %H:%M:%S').strftime('%B %d, %Y')
            cycle_info['current_close_date'] = datetime.strptime(
            str(current_cycle.bidding_close_date), '%Y-%m-%d %H:%M:%S').strftime('%B %d, %Y')

            cycle_info['open'] = True
            cycle_info['viewable'] = True

            print(cycle_info)
            return cycle_info
    except: # there is no open bidding cycle
        try: # seeing if the bidding cycle is closed but not archived
            current_cycle = UmEcommMbidCriticalDate.objects.get(bidding_closed='C')
            cycle_info['open'] = False
            cycle_info['viewable'] = True

            print(cycle_info)
            return cycle_info
        except:
            cycle_info['open'] = False
            cycle_info['viewable'] = False

            print(cycle_info)
            return cycle_info

# All MBid users views


@login_required
def home(request):
    cycle_info = cycle()
    context = {
        'title': 'MBid',
        'cycle_info': cycle_info,
        'procurement': UmBomProcurementUsersV.objects.filter(username=request.user.username.upper(), security_role_code='UM Procurement').exists(),
    }
    return render(request, 'mbid/home.html', context)


@login_required
def faq(request):
    return render(request, 'mbid/faq.html', {'title': 'FAQ'})


def complete(request, message):
    context = {'message': ''}
    if (message == 1):
        cycle_info = cycle()
        context['message'] = "Created Bid Cycle " + \
            cycle_info['current_month']+' '+cycle_info['current_year']
    if (message == 2):
        context['message'] = 'Report will be emailed to you when complete. It may take about 5 minutes.'
    return render(request, 'mbid/message.html', context)


# Mike-only views
@login_required
@permission_required('mbid_procurement', raise_exception=True)
def create_cycle(request):
    cycles = []
    for item in UmEcommMbidCriticalDate.objects.all().order_by('-bidding_year', 'bidding_month'):
        month = convert[item.bidding_month]
        cycles.append(str(month)+' '+str(item.bidding_year))

    if (request.method == 'POST') and (request.POST.get('makeCycle') == 'Create'):
        print(request.POST)
        openDateTime = request.POST.get('openDate')+' 00:00:00'
        closeDateTime = request.POST.get('closeDate')+' 23:59:59'
        print('creating critical dates')
        UmEcommMbidCriticalDate.objects.create(bidding_open_date=openDateTime,
                                               bidding_open_time=openDateTime,
                                               bidding_close_date=closeDateTime,
                                               bidding_close_time=closeDateTime,
                                               bidding_closed='O',
                                               bidding_year=request.POST.get(
                                                   'bidYear'),
                                               bidding_month=reverse_convert[request.POST.get(
                                                   'bidMonth')],
                                               date_created=datetime.now().strftime("%Y-%m-%d %H:%M"),
                                               date_last_updated=datetime.now().strftime("%Y-%m-%d %H:%M"))
        print('running proc')
        run_proc()
        print('proc complete')
        rows = []
        datalist = UmEcommMbidWarehseInput.objects.filter(bidding_closed='O')
        for data in datalist:
            umcode = data.item_code
            desc = data.item_desc
            bid_status = data.bid_status
            req_qty = data.qty_required
            annual_qty = data.annual_qty
            um_notes = data.um_notes

            data = {'umcode': umcode, 'desc': desc,
                    'req_qty': req_qty, 'annual_qty': annual_qty, 'bid_status': bid_status, 'um_notes': um_notes}
            rows.append(data)

        return render(request, 'mbid/a_createCycle.html', {
            'title': 'Create New Bid Cycle',
            'bidYear': request.POST.get('bidYear'),
            'bidMonth': request.POST.get('bidMonth'),
            'openDate': datetime.strptime(request.POST.get('openDate'), '%Y-%m-%d').strftime('%B %d, %Y'),
            'closeDate': datetime.strptime(request.POST.get('closeDate'), '%Y-%m-%d').strftime('%B %d, %Y'),
            'rows': rows,
            'totalItems': len(rows),
            'confirm': True,
        })
    else:
        form = createCycle()
        return render(request, 'mbid/a_createCycle.html', {
            'form': form,
            'title': 'Create New Bid Cycle',
            'create': True,
            'cycles': cycles
        })


@login_required
@permission_required('mbid_procurement', raise_exception=True)
def edit_cycle(request):
    cycles = []
    form = createCycle()
    cycle_info = cycle()
    for item in UmEcommMbidCriticalDate.objects.all().order_by('-bidding_year', 'bidding_month'):
        if item.bidding_closed == 'C':
            status = 'Closed'
        elif item.bidding_closed == 'O':
            status = 'Open/Current'
        else:
            status = 'Archived'
        month = convert[item.bidding_month]
        cycles.append({'year': item.bidding_year, 'month': month, 'info': str(
            item.bidding_month)+' '+str(item.bidding_year), 'status': status})

    if (request.method == 'POST'):
        target = request.POST.get('info').split()
        month = target[0]
        year = target[1]
        result = UmEcommMbidCriticalDate.objects.get(bidding_month=month, bidding_year=year)
        if request.POST.get('cycleAction') == 'edit':
            oldOpenDate = result.bidding_open_date.strftime("%Y-%m-%d")
            oldCloseDate = result.bidding_close_date.strftime("%Y-%m-%d")
            newOpenDate = request.POST.get('newOpenDate')
            newCloseDate = request.POST.get('newCloseDate')

            if (newOpenDate != ''):
                UmEcommMbidCriticalDate.objects.filter(bidding_month = month, bidding_year = year).update(bidding_open_date = newOpenDate+' 00:00:00', bidding_open_time=newOpenDate+' 00:00:00')
                # result.bidding_open_time = newOpenDate+' 00:00:00'
                # result.bidding_open_date = newOpenDate+' 00:00:00'
                # result.save()
            # elif (newOpenDate == ''):
            #     UmEcommMbidCriticalDate.objects.filter(bidding_month = month, bidding_year = year).update(bidding_open_time=newOpenDate+' 00:00:00', bidding_open_date = newOpenDate+' 00:00:00')
            #     newOpenDate = oldOpenDate
            if (newCloseDate != ''):
                UmEcommMbidCriticalDate.objects.filter(bidding_month = month, bidding_year = year).update(bidding_close_date = newCloseDate+' 00:00:00', bidding_close_time=newCloseDate+' 00:00:00')
                # result.bidding_close_time = newCloseDate+' 23:59:59'
                # result.bidding_close_date = newCloseDate+' 23:59:59'
                # result.save()
            elif (newCloseDate == ''):
                newCloseDate = oldCloseDate
            message = "Bid Cycle dates changed."
        elif request.POST.get('cycleAction') == 'close':
            earlydate = datetime.now().strftime("%Y-%m-%d %H:%M")
            UmEcommMbidCriticalDate.objects.filter(bidding_month=month, bidding_year=year).update(bidding_closed='C', bidding_close_date = earlydate, bidding_close_time = earlydate)
            
            # result.bidding_close_date = datetime.now().strftime("%Y-%m-%d %H:%M")
            # result.bidding_close_time = datetime.now().strftime("%Y-%m-%d %H:%M")
            # result.save()

            message = 'Bidding cycle {} {} has been closed.'.format(convert[month], str(year))
            # run_proc()
            # UmEcommMbidVendorInput.objects.filter(
            #     bidding_month=month, bidding_year=year).update(bidding_closed='C')

        elif request.POST.get('cycleAction') == 'archive':
            UmEcommMbidCriticalDate.objects.filter(bidding_month=month, bidding_year=year).update(bidding_closed='A')
            # result.bidding_closed = 'A'
            # result.date_last_updated = datetime.now().strftime("%Y-%m-%d %H:%M")
            # result.save()
            message = 'Bidding cycle {} {} has been archived.'.format(convert[month], str(year))
            # run_proc()
            # UmEcommMbidVendorInput.objects.filter(
            #     bidding_month=month, bidding_year=year).update(bidding_closed='A')
        elif request.POST.get('cycleAction') == 'delete':
            UmEcommMbidCriticalDate.objects.get(bidding_month=month, bidding_year=year).delete()
            UmEcommMbidWarehseInput.objects.filter(bidding_month=month, bidding_year=year).delete()
            message= 'deleted'

        return render(request, 'mbid/message.html', context={'message': message})

    rows = []
    if (cycle_info['open'] == True):
        datalist = UmEcommMbidWarehseInput.objects.filter(bidding_closed='O')
        for data in datalist:
            umcode = data.item_code
            desc = data.item_desc
            bid_status = data.bid_status
            req_qty = data.qty_required
            annual_qty = data.annual_qty
            um_notes = data.um_notes

            data = {'umcode': umcode, 'desc': desc,
                    'req_qty': req_qty, 'annual_qty': annual_qty, 'bid_status': bid_status, 'um_notes': um_notes}
            rows.append(data)

    return render(request, 'mbid/a_editCycle.html', {
        'title': 'Edit Bid Cycle',
        'cycles': cycles,
        'cycle_info': cycle_info,
        'rows': rows,
        'totalItems': len(rows)
    })


@login_required
@permission_required('mbid_procurement', raise_exception=True)
def review(request):
    # Get available cycles
    cycles = []
    for item in UmEcommMbidCriticalDate.objects.all().order_by('-bidding_year', 'bidding_month'):
        if item.bidding_closed == 'C':
            status = 'Closed'
        elif item.bidding_closed == 'A':
            status = 'Archived'
        elif item.bidding_closed == 'O':
            status = 'Open'
        else:
            status = 'idk what '+str(item.bidding_closed)+' means'
        month = convert[item.bidding_month]
        cycles.append((item.bidding_year, month, status, item.bidding_month))

    if request.method == 'POST':
        # return create_mike_report(request)
        thread = threading.Thread(
            target=create_mike_report, args=(request,))
        thread.start()
        return redirect('complete/2')
        
        
    return render(request, 'mbid/a_reviewBids.html', {'title': 'Review Bid Cycles', 'cycles': cycles})


# Vendor-only views
@login_required
@permission_required('is_vendor', raise_exception=True)
def create_cycle_report(request):
    cycle_info = cycle()
    if (cycle_info['open'] == True):
        # Use CommodityV because it's supposed to pull from current cycle
        # and includes manufacturer_part_number
        datalist = UmEcommMbidCommodityV.objects.filter(
            bidding_year=cycle_info['current_year']).order_by('manufacturer_id')
    else:
        return render(request, 'mbid/message.html', {'message': 'The bid cycle is closed and you may not place bids.'})

    rows = []

    # Prep to create workbook
    wb = Workbook()
    ws = wb.active
    
    fieldnames = ['Manufacturer', 'UM Code', 'Manufacturer Part Number', 'Description', 'Bid Status',
                  'Required QTY @ Local Vendor Branch', 'Estimated Annual Qty', 'Unit of Measure', 'UM Notes', 'Vendor Notes', 'Bid Price']

    ws.append(fieldnames)

    for data in datalist:
        mfr = data.manufacturer_id
        umcode = data.item_code
        desc = data.title
        manufacturer_part_number = data.manufacturer_part_number
        bid_status = data.bid_status
        req_qty = data.qty_required
        annual_qty = data.annual_qty
        uom = data.uom
        um_notes = data.um_notes
        rows.append({'mfr': mfr, 'umcode': umcode, 'manufacturer_part_number': manufacturer_part_number, 'desc': desc, 'bid_status': bid_status,
                     'req_qty': req_qty, 'annual_qty': annual_qty, 'uom': uom, 'um_notes': um_notes})
        ws.append([mfr, umcode, manufacturer_part_number, desc, bid_status, req_qty, annual_qty, uom, um_notes])
       
    # Only allow vendor notes and bid price columns to be edited
    ws.protection.sheet = True
    for cell in ws['J']:
        cell.protection = Protection(locked=False)
    for cell in ws['K']:
        cell.protection = Protection(locked=False)

    # Sheet formatting widths to auto
    ws.freeze_panes = ws['A2']
    for letter in ['A','B','C','D','E','F','G','H','I','J','K']:
        ws.column_dimensions[letter].auto_size = True

    if request.method == 'POST':
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="MBid_' + \
            str(cycle_info['current_month'])+'_' + \
            str(cycle_info['current_year'])+'.xlsx"'
        return response

    context = {
        'title': 'Create Report',
        'rows': rows,
        'totalItems': len(rows),
        'cycle_info': cycle_info
    }
    return render(request, 'mbid/c_cycleExport.html', context)


@login_required
@permission_required('is_vendor', raise_exception=True)
def upload_bids(request):
    # Two parts to this view
    # 1. allowing vendors to upload
    # 2. reviewing the vendor upload before submitting to databse
    cycle_info = cycle()
    context = {'title': 'Upload Bids', 'cycle_info': cycle_info}
    try: # No matter who it is you NEED to be in the vendor table to upload
        vendor = UmEcommMbidVendorV.objects.get(
            email_address=request.user.email)
    except:
        return render(request, 'mbid/message.html', {'message': 'You do not have permission to view this'})

    # Vendor bids get uploaded if there is an open cycle, and the datetime is within the bid cycle
    if ((request.method == "POST") and 
    (UmEcommMbidCriticalDate.objects.filter(bidding_closed='O').exists()) and 
    (datetime.now() > UmEcommMbidCriticalDate.objects.get(bidding_closed='O').bidding_open_time) and 
    (datetime.now() < UmEcommMbidCriticalDate.objects.get(bidding_closed='O').bidding_close_time) and 
    (request.POST.get('uploadstatus') == 'Upload Bids')):
        file_in_memory = request.FILES['mbid_file'].read()
        wb = load_workbook(filename=BytesIO(file_in_memory))
        # wb = load_workbook(filename=request.FILES['mbid_file'].file)
        worksheet = wb[wb.sheetnames[0]]
        reader=list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            row = ['' if cell==None else cell for cell in row]
            reader.append({
                'Manufacturer':row[0],
                'UMCode':row[1], 
                'ManufacturerPartNumber':row[2],
                'Description':row[3],
                'BidStatus':row[4],
                'RequiredQTY':row[5],
                'AnnualQty':row[6], 
                'UnitofMeasure':row[7], 
                'UMNotes':row[8], 
                'VendorNotes':row[9], 
                'BidPrice':row[10]})
        
        # data validation
        valid = []
        notvalid = []

        for row in reader:

            if (pricecheck(row['BidPrice']) != None) and (pricecheck(row['BidPrice']) != False):
                valid.append(row)
            # If empty and vendor notes
            elif (pricecheck(row['BidPrice']) == None) and (row['VendorNotes'] != ''):
                notvalid.append(row)
            # If there's something in the field and it's an invalid price (0 or string)
            elif (pricecheck(row['BidPrice']) == False):
                notvalid.append(row)
            # If for some reason none of the above, ignore it
            else:
                pass

        mistakes = False
        if (notvalid != []):
            mistakes = True
        context = {
            'valid': valid,
            'mistakes': mistakes,
            'notvalid': notvalid,
            'title': 'Uploaded Bids'}

        return render(request, 'mbid/c_checkLoad.html', context)
    elif (request.method == 'POST') and (request.POST.get('uploadstatus') == 'Submit Bids'):
        thread = threading.Thread(target=update_vendor_table, args=(request,))
        thread.start()
        # EDIT MESSAGE 
        return render(request, 'mbid/message.html', {'title': 'Upload Bids', 'message': 'Your bids are being uploaded.'})
    return render(request, 'mbid/c_uploadBids.html', context)


@login_required
@permission_required('is_vendor', raise_exception=True)
def review_bids(request):
    cycle_info = cycle()
    if (cycle_info['open'] == True) or (cycle_info['viewable'] == True):
        try:
            inputlist = UmEcommMbidVendorInput.objects.filter(
                bidding_year=cycle_info['current_year'],
                bidding_month=cycle_info['current_month_num'],
                vendor_id=UmEcommMbidVendorV.objects.get(email_address=request.user.email).id)
        except:
            return render(request, 'mbid/message.html', {'message': 'Vendor not found'})

    else:
        return render(request, 'mbid/message.html', {'message': 'The bid cycle is closed and you can no longer view your bids.'})

    rows = []

    for item in inputlist:
        extra = UmEcommMbidCommodityV.objects.get(
            bidding_year=cycle_info['current_year'], bidding_month=cycle_info['current_month_num'], item_code=item.item_code)

        manufacturer_name = item.manufacturer_name
        umcode = item.item_code
        desc = item.item_desc
        vendor_notes = item.vendor_notes
        price = item.vendor_price
        manufacturer_part_number = item.manufacturer_part_number
        uom = extra.uom
        req_qty = extra.qty_required
        annual_qty = extra.annual_qty
        bid_status = extra.bid_status
        um_notes = extra.um_notes

        data = {'manufacturer_name': manufacturer_name, 'umcode': umcode, 'desc': desc, 'manufacturer_part_number': manufacturer_part_number, 'bid_status': bid_status,
                'um_notes': um_notes, 'uom': uom, 'req_qty': req_qty, 'annual_qty': annual_qty, 'vendor_notes': vendor_notes, 'price': price}
        rows.append(data)

    context = {
        'title': 'Review Bids',
        'rows': [],
        'rows': rows,
        'totalItems': len(rows),
        'noitems': (len(rows) == 0),
        'cycle_info': cycle_info
    }
    return render(request, 'mbid/c_cycleReport.html', context)

# Not views

def pricecheck(bidprice):
    # return None if can't be converted/empty
    if (bidprice == ''):
        return None
    
    # return False if invalid, return float if ok
    try:
        num = float(bidprice)
        if (num >= 0) and (num == round(num, 2)): 
            return str(round(num, 2)) # standardize bid to 2 decimal points
        else:
            return False # invalid due to being less than 0
    except:
        return False   # invalid due to being a string


def run_proc():
    curr = connections['pinnacle'].cursor()
    try:
        curr.callproc(
            'UM_VENDOR_BIDDING_INTERFACE_K.UM_UPDATE_WAREHSE_INPUT_P')
    except:
        print('error')
    curr.close()

# # Send emails via thread

def update_vendor_table(request):
    cycle_info = cycle()
    post = request.POST
    vendor = UmEcommMbidVendorV.objects.get(email_address=request.user.email)
    vendor_id = vendor.id
    vendor_name = vendor.name
    vendor_address1 = vendor.address1
    vendor_address2 = vendor.address2
    vendor_city = vendor.city
    vendor_state = vendor.state
    vendor_zip_code = vendor.zip_code  # no zip extension
    vendor_email_address = vendor.email_address
    havetarget = False
    
    # clear previous bids if any
    UmEcommMbidVendorInput.objects.filter(
                bidding_year=cycle_info['current_year'],
                bidding_month=cycle_info['current_month_num'],
                vendor_id=vendor_id).delete()
    
    # uses the previously validated list bids
    for string in post.getlist('uploads'):
        row = dict({x.split(':')[0]: x.split(':')[1]
                    for x in [item for item in string.split(';')]})
        UmEcommMbidVendorInput.objects.create(
            bidding_year=cycle_info['current_year'],
            bidding_month=cycle_info['current_month_num'],
            bidding_closed='O',
            item_code=row['UMCode'],
            item_desc=row['Description'],
            subclass_id=UmEcommMbidCommodityV.objects.get(
                item_code=row['UMCode']).subclass_id,
            manufacturer_name=row['Manufacturer'],
            manufacturer_part_number=row['ManufacturerPartNumber'],
            vendor_id=vendor_id,
            vendor_name=vendor_name,
            vendor_email_address=vendor_email_address,
            vendor_address1=vendor_address1,
            vendor_address2=vendor_address2,
            vendor_city=vendor_city,
            vendor_state=vendor_state,
            vendor_zip_code=vendor_zip_code,
            vendor_price=float(row['BidPrice']),
            vendor_notes=str(row['VendorNotes']),
            date_created=datetime.now().strftime("%Y-%m-%d %H:%M"),
            date_last_updated=datetime.now().strftime("%Y-%m-%d %H:%M")
            )

    print('complete upload')
    email = EmailMessage(
        subject='MBid Update',
        body='Your bids were uploaded',
        from_email='srs@umich.edu',
        to=[request.user.email],
    )
    email.send()

def create_mike_report(request):
    # Make sure post is a dictionary for use in determining download option
    # 'pickCycle': '02 2023', 'downloadOption': 'allBids'
    post = request.POST.dict()

    # Ready email
    email = EmailMessage(
        subject='MBid Report',
        body='See attached',
        from_email='srs@umich.edu',
        to=[request.user.email],)

    # get month/year for queries
    yearmonth = post['pickCycle'].split()
    bidding_month = yearmonth[0]
    bidding_year = yearmonth[1]

    # Set up Excel
    target_xlsx = BytesIO()
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    # Retrieve data as specified by selected option
    # If no vendor notes or bids for noBids option
    if (post['downloadOption'] == 'noBids'):
        # Set up header
        fieldnames = ['U-M Code', 'Description', 'Bid Status', 'UM Notes']
        ws.append(fieldnames)

        # Get data
        havebids = set(UmEcommMbidVendorInput.objects.filter(
            bidding_year=bidding_year, bidding_month=bidding_month).values_list('item_code', flat=True))
        rows = UmEcommMbidWarehseInput.objects.filter(
            bidding_year=bidding_year, bidding_month=bidding_month).exclude(item_code__in=havebids)
        
        # Write data, no vendor info since these have no bids
        for row in rows:
            ws.append([row.item_code, row.item_desc, row.bid_status, row.um_notes])

        # for actual
        wb.save(target_xlsx)
        email.attach('MikeReport_' + str(bidding_month)+'_'+str(bidding_year) + '_noBids.xlsx', target_xlsx.getvalue(), 'application/ms-excel')

        # for testing
        # response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        # response['Content-Disposition'] = 'attachment; filename="MikeReport_' + \
        #     str(convert[bidding_month])+'_' + \
        #     str(bidding_year)+'_noBids.xlsx"'
        # return response


    elif (post['downloadOption'] == 'allBids'):
        # Set up header
        fieldnames = ['U-M Code', 'Description', 'Bid Status']
        # Add all vendors that bidded in cycle to header
        for vendor in set(UmEcommMbidVendorInput.objects.filter(bidding_year=bidding_year,bidding_month=bidding_month).values_list('vendor_id', flat=True)):
            fieldnames.extend([vendor + ' Notes', vendor + ' Bids'])

        ws.append(fieldnames)

        # Get list of UM Codes with bids
        item_codes_list = sorted(set(UmEcommMbidVendorInput.objects.filter(
            bidding_year=bidding_year, bidding_month=bidding_month).values_list('item_code', flat=True)))

        # Write data, each row is item, each item add vendor info (to dict), convert to simple list in end for appending
        for item in item_codes_list:
            # convert fieldnames to dictionary for easy locating
            fieldnamesDict = OrderedDict((field, '') for field in fieldnames)
            
            info = UmEcommMbidWarehseInput.objects.get(item_code=item, bidding_year=bidding_year, bidding_month=bidding_month)
            fieldnamesDict['U-M Code'] = item
            fieldnamesDict['Description'] = info.item_desc
            fieldnamesDict['Bid Status'] = info.bid_status

            results = UmEcommMbidVendorInput.objects.filter(bidding_year=bidding_year, bidding_month=bidding_month, item_code=item)
            for row in results:
                fieldnamesDict[row.vendor_id+' Notes'] = row.vendor_notes
                fieldnamesDict[row.vendor_id + ' Bids'] = row.vendor_price

            ws.append(list(fieldnamesDict.values()))

        # for actual
        wb.save(target_xlsx)
        email.attach('MikeReport_' + str(convert[bidding_month])+'_'+str(bidding_year) + '_allBids.xlsx', target_xlsx.getvalue(), 'application/ms-excel')

        # for testing
        # response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        # response['Content-Disposition'] = 'attachment; filename="MikeReport_' + \
        #     str(convert[bidding_month])+'_' + \
        #     str(bidding_year)+'_allBids.xlsx"'
        # return response

    elif post.get('downloadOption') == 'lowBids':
        # Set header for 3 lowest bidders
        fieldnames = ['U-M Code', 'Description', 'Bid Status', 'Vendor 1', 'Vendor Notes 1', 'Vendor Price 1',
                      'Vendor 2', 'Vendor Notes 2', 'Vendor Price 2', 'Vendor 3', 'Vendor Notes 3', 'Vendor Price 3']
        ws.append(fieldnames)

        # Get list of UM Codes with bids
        item_codes_list = sorted(set(UmEcommMbidVendorInput.objects.filter(
            bidding_year=bidding_year, bidding_month=bidding_month).values_list('item_code', flat=True)))

        # Write data
        for item in item_codes_list:
            # Get vendor id,notes,bids for item. order by price
            results = UmEcommMbidVendorInput.objects.filter(
                bidding_year=bidding_year, bidding_month=bidding_month, item_code=item).order_by('vendor_price')[:3]
            
            # get bid status from warehouse
            info = UmEcommMbidWarehseInput.objects.get(
                item_code=item, bidding_year=bidding_year, bidding_month=bidding_month)
           
            towrite = [info.item_code, info.item_desc, info.bid_status]  

            if len(results) == 3:
                towrite.extend([results[0].vendor_id, results[0].vendor_notes, results[0].vendor_price,
                results[1].vendor_id, results[1].vendor_notes, results[1].vendor_price,
                results[2].vendor_id, results[2].vendor_notes, results[2].vendor_price])

            elif len(results) == 2:
                towrite.extend([results[0].vendor_id,results[0].vendor_notes,results[0].vendor_price,
                results[1].vendor_id,results[1].vendor_notes,results[1].vendor_price])

            elif len(results) == 1:
                towrite.extend([results[0].vendor_id, results[0].vendor_notes, results[0].vendor_price])

            ws.append(towrite)

        # for actual
        wb.save(target_xlsx)
        email.attach('MikeReport_' + str(bidding_month)+'_'+str(bidding_year) + '_lowestBids.xlsx', target_xlsx.getvalue(), 'application/ms-excel')
        
        # for testing
        # response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        # response['Content-Disposition'] = 'attachment; filename="MikeReport_' + \
        #     str(convert[bidding_month])+'_' + \
        #     str(bidding_year)+'_lowBids.xlsx"'
        # return response

    # Everything written and attached in if statements. Send email.
    email.send(fail_silently=False)
